#!/usr/bin/env python3
"""
STARFORGE TCG — v2 Card Converter
Converts v2 Excel card files into:
  1. src/data/SampleCards.ts (full TypeScript with typed effects)
  2. src/data/BalancedStarterDecks.ts (30-card starter decks per faction)
  3. art-export/{Faction}/cards.json (updated JSON for art pipeline)
  4. art-export/all-cards.json (combined)
"""

import openpyxl
import os
import re
import json
from collections import defaultdict

CARDS_DIR = "/home/user/starforge_tcg/new-cards"
SRC_DIR = "/home/user/starforge_tcg/src/data"
ART_DIR = "/home/user/starforge_tcg/art-export"

# Map faction names to Race enum values and prefixes
FACTION_MAP = {
    "Astromancers": {"race": "Race.ASTROMANCERS", "prefix": "ast", "enum": "ASTROMANCERS"},
    "Biotitans": {"race": "Race.BIOTITANS", "prefix": "bio", "enum": "BIOTITANS"},
    "Chronobound": {"race": "Race.CHRONOBOUND", "prefix": "chr", "enum": "CHRONOBOUND"},
    "Cogsmiths": {"race": "Race.COGSMITHS", "prefix": "cog", "enum": "COGSMITHS"},
    "Crystalline": {"race": "Race.CRYSTALLINE", "prefix": "cry", "enum": "CRYSTALLINE"},
    "Hivemind": {"race": "Race.HIVEMIND", "prefix": "hiv", "enum": "HIVEMIND"},
    "Luminar": {"race": "Race.LUMINAR", "prefix": "lum", "enum": "LUMINAR"},
    "PhantomCorsairs": {"race": "Race.PHANTOM_CORSAIRS", "prefix": "phc", "enum": "PHANTOM_CORSAIRS"},
    "Pyroclast": {"race": "Race.PYROCLAST", "prefix": "pyr", "enum": "PYROCLAST"},
    "Voidborn": {"race": "Race.VOIDBORN", "prefix": "voi", "enum": "VOIDBORN"},
}

# Keyword mapping to TypeScript enums
KEYWORD_MAP = {
    "GUARDIAN": "CombatKeyword.GUARDIAN",
    "BARRIER": "CombatKeyword.BARRIER",
    "SWIFT": "CombatKeyword.SWIFT",
    "BLITZ": "CombatKeyword.BLITZ",
    "CLOAK": "CombatKeyword.CLOAK",
    "DOUBLE STRIKE": "CombatKeyword.DOUBLE_STRIKE",
    "DOUBLE_STRIKE": "CombatKeyword.DOUBLE_STRIKE",
    "DRAIN": "CombatKeyword.DRAIN",
    "LETHAL": "CombatKeyword.LETHAL",
    "DEPLOY": "TriggerKeyword.DEPLOY",
    "LAST WORDS": "TriggerKeyword.LAST_WORDS",
    "LAST_WORDS": "TriggerKeyword.LAST_WORDS",
    "SALVAGE": "OriginalKeyword.SALVAGE",
    "UPGRADE": "OriginalKeyword.UPGRADE",
    "ILLUMINATE": "OriginalKeyword.ILLUMINATE",
    "IMMOLATE": "OriginalKeyword.IMMOLATE",
    "BANISH": "OriginalKeyword.BANISH",
    "ADAPT": "OriginalKeyword.ADAPT",
    "RESONATE": "OriginalKeyword.RESONATE",
    "PHASE": "OriginalKeyword.PHASE",
    "SWARM": "OriginalKeyword.SWARM",
    "SCRY": "OriginalKeyword.SCRY",
    "ECHO": "OriginalKeyword.ECHO",
}

TRIBE_MAP = {
    "MECH": "MinionTribe.MECH",
    "BEAST": "MinionTribe.BEAST",
    "ELEMENTAL": "MinionTribe.ELEMENTAL",
    "DRAGON": "MinionTribe.DRAGON",
    "PIRATE": "MinionTribe.PIRATE",
    "DEMON": "MinionTribe.DEMON",
    "INSECT": "MinionTribe.INSECT",
    "CONSTRUCT": "MinionTribe.CONSTRUCT",
    "VOID": "MinionTribe.VOID",
}

RARITY_MAP = {
    "COMMON": "CardRarity.COMMON",
    "RARE": "CardRarity.RARE",
    "EPIC": "CardRarity.EPIC",
    "LEGENDARY": "CardRarity.LEGENDARY",
}

TYPE_MAP = {
    "MINION": "CardType.MINION",
    "SPELL": "CardType.SPELL",
    "STRUCTURE": "CardType.STRUCTURE",
}


def load_excel(path):
    wb = openpyxl.load_workbook(path, read_only=True)
    sheets = {}
    for sname in wb.sheetnames:
        ws = wb[sname]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = rows[0]
        cards = [dict(zip(headers, r)) for r in rows[1:] if any(v is not None for v in r)]
        sheets[sname] = cards
    wb.close()
    return sheets


def parse_keywords(kw_str):
    """Parse keyword string into list of (keyword_name, optional_value)."""
    if not kw_str or str(kw_str).strip() == "":
        return []
    results = []
    for part in str(kw_str).split(","):
        part = part.strip().upper()
        if not part:
            continue
        # Handle "SCRY (2)" or "UPGRADE (3)"
        m = re.match(r'(\w[\w\s]*?)\s*\((\d+)\)', part)
        if m:
            name = m.group(1).strip()
            val = int(m.group(2))
            results.append((name, val))
        else:
            results.append((part, None))
    return results


def make_id(prefix, num, card_type):
    type_code = {"Minion": "m", "Spell": "s", "Structure": "st"}.get(card_type, "c")
    return f"{prefix}_{type_code}{num}"


def escape_ts_string(s):
    if not s:
        return ""
    return str(s).replace("\\", "\\\\").replace("'", "\\'").replace("\n", " ")


def parse_card_text_to_effects(card_text, card_type, keywords_parsed):
    """
    Parse card text into TypeScript effect objects.
    Returns a list of effect strings for the effects[] array.
    """
    if not card_text or card_text.strip() == "" or card_text.strip() == ".":
        return []

    text = str(card_text).upper()
    effects = []
    effect_id = 0

    kw_names = [k[0] for k in keywords_parsed]
    has_deploy = "DEPLOY" in kw_names
    has_last_words = "LAST WORDS" in kw_names or "LAST_WORDS" in kw_names

    # Determine trigger
    if "DEPLOY:" in text or "DEPLOY." in text:
        trigger = "EffectTrigger.ON_PLAY"
    elif "LAST WORDS:" in text:
        trigger = "EffectTrigger.ON_DEATH"
    elif "START OF TURN" in text:
        trigger = "EffectTrigger.ON_TURN_START"
    elif "END OF TURN" in text:
        trigger = "EffectTrigger.ON_TURN_END"
    elif "RESONATE" in text:
        trigger = "EffectTrigger.ON_SPELL_CAST"
    elif "ILLUMINATE" in text:
        trigger = "EffectTrigger.ON_HEAL"
    elif "IMMOLATE" in text:
        trigger = "EffectTrigger.ON_DEATH"
    elif "WHENEVER" in text and "ATTACK" in text:
        trigger = "EffectTrigger.ON_ATTACK"
    elif "WHENEVER" in text and "SPELL" in text:
        trigger = "EffectTrigger.ON_SPELL_CAST"
    elif "WHENEVER" in text and "HEAL" in text:
        trigger = "EffectTrigger.ON_HEAL"
    elif "WHENEVER" in text and ("DIE" in text or "DIES" in text or "DESTROYED" in text):
        trigger = "EffectTrigger.ON_FRIENDLY_DEATH"
    elif "WHENEVER" in text and "SUMMON" in text:
        trigger = "EffectTrigger.ON_SUMMON"
    elif has_deploy:
        trigger = "EffectTrigger.ON_PLAY"
    elif has_last_words:
        trigger = "EffectTrigger.ON_DEATH"
    elif card_type == "Spell":
        trigger = "EffectTrigger.ON_PLAY"
    else:
        trigger = "EffectTrigger.ON_PLAY"

    orig_text = str(card_text)

    # DEAL X DAMAGE
    dmg_matches = re.findall(r'DEAL\s+(\d+)\s+DAMAGE', text)
    for dmg_str in dmg_matches:
        dmg = int(dmg_str)
        if "ALL ENEMIES" in text or "ALL ENEMY MINIONS" in text:
            target = "TargetType.ALL_ENEMY_MINIONS"
        elif "ALL MINIONS" in text:
            target = "TargetType.ALL_MINIONS"
        elif "RANDOM ENEMY" in text:
            target = "TargetType.RANDOM_ENEMY"
        elif "ENEMY HERO" in text:
            target = "TargetType.ENEMY_HERO"
        else:
            target = "TargetType.RANDOM_ENEMY"
        effects.append(
            f"{{ id: 'e{effect_id}', type: EffectType.DAMAGE, trigger: {trigger}, "
            f"targetType: {target}, data: {{ amount: {dmg} }} as DamageEffectData, isMandatory: true }}"
        )
        effect_id += 1

    # DRAW X CARD(S)
    draw_matches = re.findall(r'DRAW\s+(\d+|A)\s+CARD', text)
    if not draw_matches and "DRAW A CARD" in text:
        draw_matches = ["1"]
    if not draw_matches and re.search(r'DRAW\s+A\s+CARD', text):
        draw_matches = ["1"]
    for draw_str in draw_matches:
        count = 1 if draw_str == "A" else int(draw_str)
        effects.append(
            f"{{ id: 'e{effect_id}', type: EffectType.DRAW, trigger: {trigger}, "
            f"targetType: TargetType.NONE, data: {{ count: {count} }} as DrawEffectData, isMandatory: true }}"
        )
        effect_id += 1

    # Simple "DRAW A CARD" without number
    if not draw_matches and "DRAW A CARD" in text:
        effects.append(
            f"{{ id: 'e{effect_id}', type: EffectType.DRAW, trigger: {trigger}, "
            f"targetType: TargetType.NONE, data: {{ count: 1 }} as DrawEffectData, isMandatory: true }}"
        )
        effect_id += 1

    # HEAL / RESTORE X HEALTH
    heal_matches = re.findall(r'(?:HEAL|RESTORE)\s+(\d+)\s+HEALTH', text)
    for heal_str in heal_matches:
        heal = int(heal_str)
        if "ALL FRIENDLY" in text or "ALL FRIENDLIES" in text:
            target = "TargetType.ALL_FRIENDLIES"
        elif "YOUR HERO" in text or "FRIENDLY HERO" in text:
            target = "TargetType.FRIENDLY_HERO"
        else:
            target = "TargetType.CHOSEN"
        effects.append(
            f"{{ id: 'e{effect_id}', type: EffectType.HEAL, trigger: {trigger}, "
            f"targetType: {target}, data: {{ amount: {heal} }} as HealEffectData, isMandatory: true }}"
        )
        effect_id += 1

    # BUFF: GIVE +X/+Y or GAIN +X/+Y
    buff_matches = re.findall(r'(?:GIVE|GAIN|GET)\s+\+(\d+)/\+(\d+)', text)
    for atk_str, hp_str in buff_matches:
        atk, hp = int(atk_str), int(hp_str)
        if "ALL FRIENDLY MINIONS" in text or "ALL FRIENDLY" in text:
            target = "TargetType.ALL_FRIENDLY_MINIONS"
        elif "ALL MINIONS" in text:
            target = "TargetType.ALL_MINIONS"
        elif "GAIN" in text:
            target = "TargetType.SELF"
        else:
            target = "TargetType.CHOSEN"
        effects.append(
            f"{{ id: 'e{effect_id}', type: EffectType.BUFF, trigger: {trigger}, "
            f"targetType: {target}, data: {{ attack: {atk}, health: {hp} }} as BuffEffectData, isMandatory: true }}"
        )
        effect_id += 1

    # BUFF: +X ATTACK only
    atk_only = re.findall(r'(?:GIVE|GAIN)\s+\+(\d+)\s+ATTACK', text)
    for atk_str in atk_only:
        atk = int(atk_str)
        target = "TargetType.SELF" if "GAIN" in text else "TargetType.CHOSEN"
        effects.append(
            f"{{ id: 'e{effect_id}', type: EffectType.BUFF, trigger: {trigger}, "
            f"targetType: {target}, data: {{ attack: {atk}, health: 0 }} as BuffEffectData, isMandatory: true }}"
        )
        effect_id += 1

    # BUFF: +X HEALTH only
    hp_only = re.findall(r'(?:GIVE|GAIN)\s+\+(\d+)\s+HEALTH', text)
    for hp_str in hp_only:
        hp = int(hp_str)
        target = "TargetType.SELF" if "GAIN" in text else "TargetType.CHOSEN"
        effects.append(
            f"{{ id: 'e{effect_id}', type: EffectType.BUFF, trigger: {trigger}, "
            f"targetType: {target}, data: {{ attack: 0, health: {hp} }} as BuffEffectData, isMandatory: true }}"
        )
        effect_id += 1

    # SUMMON
    summon_matches = re.findall(r'SUMMON\s+(?:A\s+|TWO\s+|THREE\s+|FOUR\s+|(\d+)\s+)?(\d+)/(\d+)', text)
    if not summon_matches:
        summon_count_match = re.findall(r'SUMMON\s+(TWO|THREE|FOUR|\d+)', text)
        if summon_count_match:
            word = summon_count_match[0]
            count = {"TWO": 2, "THREE": 3, "FOUR": 4}.get(word, 1)
            if count == 1:
                try:
                    count = int(word)
                except:
                    count = 1
            effects.append(
                f"{{ id: 'e{effect_id}', type: EffectType.SUMMON, trigger: {trigger}, "
                f"targetType: TargetType.NONE, data: {{ cardId: 'token_generated', count: {count} }} as SummonEffectData, isMandatory: true }}"
            )
            effect_id += 1
        elif "SUMMON" in text and "DRONE" in text:
            count_match = re.search(r'(\d+|TWO|THREE)\s+.*DRONE', text)
            count = 2
            if count_match:
                word = count_match.group(1)
                count = {"TWO": 2, "THREE": 3}.get(word, 2)
                try:
                    count = int(word)
                except:
                    pass
            effects.append(
                f"{{ id: 'e{effect_id}', type: EffectType.SUMMON, trigger: {trigger}, "
                f"targetType: TargetType.NONE, data: {{ cardId: 'token_drone', count: {count} }} as SummonEffectData, isMandatory: true }}"
            )
            effect_id += 1
    else:
        for count_str, atk_str, hp_str in summon_matches:
            count = 1
            # Check for word-counts before the stats
            pre_text = text.split(f"{atk_str}/{hp_str}")[0] if f"{atk_str}/{hp_str}" in text else text
            if "TWO" in pre_text:
                count = 2
            elif "THREE" in pre_text:
                count = 3
            elif count_str:
                count = int(count_str)
            effects.append(
                f"{{ id: 'e{effect_id}', type: EffectType.SUMMON, trigger: {trigger}, "
                f"targetType: TargetType.NONE, data: {{ cardId: 'token_generated', count: {count} }} as SummonEffectData, isMandatory: true }}"
            )
            effect_id += 1

    # DESTROY a minion
    if "DESTROY A MINION" in text or "DESTROY AN ENEMY MINION" in text:
        target = "TargetType.ENEMY_MINION" if "ENEMY" in text else "TargetType.CHOSEN"
        effects.append(
            f"{{ id: 'e{effect_id}', type: EffectType.DESTROY, trigger: {trigger}, "
            f"targetType: {target}, data: {{}} as GenericEffectData, isMandatory: true }}"
        )
        effect_id += 1

    # GAIN X CRYSTAL(S)
    crystal_matches = re.findall(r'GAIN\s+(\d+)\s+(?:EMPTY\s+)?CRYSTAL', text)
    for c_str in crystal_matches:
        effects.append(
            f"{{ id: 'e{effect_id}', type: EffectType.GAIN_CRYSTALS, trigger: {trigger}, "
            f"targetType: TargetType.NONE, data: {{ amount: {int(c_str)} }} as GainCrystalsData, isMandatory: true }}"
        )
        effect_id += 1

    # RETURN TO HAND
    if "RETURN" in text and "HAND" in text:
        effects.append(
            f"{{ id: 'e{effect_id}', type: EffectType.RETURN_TO_HAND, trigger: {trigger}, "
            f"targetType: TargetType.FRIENDLY_MINION, data: {{}} as GenericEffectData, isMandatory: true }}"
        )
        effect_id += 1

    # ADAPT
    if "ADAPT" in kw_names:
        effects.append(
            f"{{ id: 'e{effect_id}', type: EffectType.ADAPT, trigger: EffectTrigger.ON_PLAY, "
            f"targetType: TargetType.SELF, data: {{ options: 3 }} as AdaptData, isMandatory: true }}"
        )
        effect_id += 1

    # SCRY
    scry_kw = [k for k in keywords_parsed if k[0] == "SCRY"]
    if scry_kw:
        val = scry_kw[0][1] or 2
        effects.append(
            f"{{ id: 'e{effect_id}', type: EffectType.SCRY, trigger: {trigger}, "
            f"targetType: TargetType.NONE, data: {{ count: {val} }} as ScryData, isMandatory: true }}"
        )
        effect_id += 1

    # BANISH effect in card text (not just keyword)
    if "BANISH" in text and "BANISH" not in kw_names:
        banish_count = 1
        bm = re.search(r'BANISH\s+(\d+|ONE|TWO|THREE)', text)
        if bm:
            word = bm.group(1)
            banish_count = {"ONE": 1, "TWO": 2, "THREE": 3}.get(word, 1)
            try:
                banish_count = int(word)
            except:
                pass
        effects.append(
            f"{{ id: 'e{effect_id}', type: EffectType.BANISH, trigger: {trigger}, "
            f"targetType: TargetType.RANDOM_ENEMY, data: {{ count: {banish_count} }} as GenericEffectData, isMandatory: true }}"
        )
        effect_id += 1

    # GRANT KEYWORD (in card text like "Give a minion BARRIER")
    grant_kws = []
    for kw_name in ["BARRIER", "GUARDIAN", "SWIFT", "LETHAL", "DRAIN", "CLOAK", "PHASE"]:
        if f"GIVE" in text and kw_name in text and kw_name not in kw_names:
            grant_kws.append(kw_name)
    if grant_kws and not buff_matches:  # Don't duplicate if already handled by buff
        kw_instances = ", ".join(f"{{ keyword: {KEYWORD_MAP[k]} }}" for k in grant_kws)
        if "ALL FRIENDLY" in text:
            target = "TargetType.ALL_FRIENDLY_MINIONS"
        else:
            target = "TargetType.CHOSEN"
        effects.append(
            f"{{ id: 'e{effect_id}', type: EffectType.GRANT_KEYWORD, trigger: {trigger}, "
            f"targetType: {target}, data: {{ keywords: [{kw_instances}] }} as GrantKeywordData, isMandatory: true }}"
        )
        effect_id += 1

    # DISCARD
    if "DISCARD" in text:
        discard_count = 1
        dm = re.search(r'DISCARD\s+(\d+|A)', text)
        if dm:
            word = dm.group(1)
            discard_count = 1 if word == "A" else int(word)
        target = "TargetType.ENEMY_HERO" if "OPPONENT" in text else "TargetType.NONE"
        effects.append(
            f"{{ id: 'e{effect_id}', type: EffectType.DISCARD, trigger: {trigger}, "
            f"targetType: {target}, data: {{ count: {discard_count} }} as GenericEffectData, isMandatory: true }}"
        )
        effect_id += 1

    # STEAL
    if "STEAL" in text or "COPY" in text:
        effects.append(
            f"{{ id: 'e{effect_id}', type: EffectType.STEAL, trigger: {trigger}, "
            f"targetType: TargetType.RANDOM_ENEMY, data: {{}} as GenericEffectData, isMandatory: true }}"
        )
        effect_id += 1

    # SILENCE
    if "SILENCE" in text:
        effects.append(
            f"{{ id: 'e{effect_id}', type: EffectType.SILENCE, trigger: {trigger}, "
            f"targetType: TargetType.CHOSEN, data: {{}} as GenericEffectData, isMandatory: true }}"
        )
        effect_id += 1

    # FREEZE
    if "FREEZE" in text:
        target = "TargetType.ALL_ENEMY_MINIONS" if "ALL" in text else "TargetType.CHOSEN"
        effects.append(
            f"{{ id: 'e{effect_id}', type: EffectType.DEBUFF, trigger: {trigger}, "
            f"targetType: {target}, data: {{ attack: 0, health: 0 }} as BuffEffectData, isMandatory: true }}"
        )
        effect_id += 1

    return effects


def build_keyword_instances(keywords_parsed):
    """Convert parsed keywords to TypeScript KeywordInstance[] string."""
    instances = []
    for name, value in keywords_parsed:
        ts_kw = KEYWORD_MAP.get(name)
        if not ts_kw:
            continue
        if value is not None:
            instances.append(f"{{ keyword: {ts_kw}, value: {value} }}")
        else:
            instances.append(f"{{ keyword: {ts_kw} }}")

    # Add DEPLOY/LAST_WORDS if mentioned in card text keywords
    return instances


def card_to_ts(card, faction_info, card_num, counters):
    """Convert a single card dict to a TypeScript CardDefinition string."""
    name = str(card.get("Card Name", "Unknown"))
    cost = int(card.get("Cost", 0))
    card_type = str(card.get("Type", "Minion"))
    subtype = card.get("Subtype")
    rarity = str(card.get("Rarity", "Common")).upper()
    atk = card.get("Atk")
    hp = card.get("HP")
    card_text = str(card.get("Card Text", "") or "")
    flavor = str(card.get("Flavor Text", "") or "")
    kw_str = card.get("Keywords") or ""
    starforge = card.get("STARFORGE")

    # Generate ID
    cid = make_id(faction_info["prefix"], card_num, card_type)

    # Parse keywords from the Keywords column
    keywords_parsed = parse_keywords(kw_str)

    # Also detect DEPLOY/LAST WORDS from card text if not in keywords
    text_upper = card_text.upper()
    kw_names = [k[0] for k in keywords_parsed]
    if ("DEPLOY:" in text_upper or "DEPLOY." in text_upper) and "DEPLOY" not in kw_names:
        keywords_parsed.append(("DEPLOY", None))
    if "LAST WORDS:" in text_upper and "LAST WORDS" not in kw_names and "LAST_WORDS" not in kw_names:
        keywords_parsed.append(("LAST WORDS", None))

    # Build keyword instances
    kw_instances = build_keyword_instances(keywords_parsed)
    kw_str_ts = "[" + ", ".join(kw_instances) + "]" if kw_instances else "[]"

    # Build effects
    effects_list = parse_card_text_to_effects(card_text, card_type, keywords_parsed)
    effects_str = "[" + ", ".join(effects_list) + "]" if effects_list else "[]"

    # Build the CardDefinition
    parts = []
    parts.append(f"id: '{cid}'")
    parts.append(f"name: '{escape_ts_string(name)}'")
    parts.append(f"cost: {cost}")
    parts.append(f"type: {TYPE_MAP.get(card_type.upper(), 'CardType.MINION')}")
    parts.append(f"race: {faction_info['race']}")
    parts.append(f"rarity: {RARITY_MAP.get(rarity, 'CardRarity.COMMON')}")

    if card_type == "Minion":
        parts.append(f"attack: {int(atk) if atk is not None else 0}")
        parts.append(f"health: {int(hp) if hp is not None else 1}")
        if subtype:
            tribe_key = str(subtype).upper()
            tribe_ts = TRIBE_MAP.get(tribe_key)
            if tribe_ts:
                parts.append(f"tribe: {tribe_ts}")

    parts.append(f"keywords: {kw_str_ts}")
    parts.append(f"effects: {effects_str}")

    # STARFORGE
    if starforge and str(starforge).strip():
        sf_text = escape_ts_string(str(starforge))
        parts.append(f"starforge: {{ type: StarforgeType.PROGRESSIVE, conditions: [], forgedForm: {{ name: \\'{name} (Forged)\\', cost: {cost}, keywords: [], effects: [], cardText: \\'{sf_text}\\' }} }}")

    parts.append(f"flavorText: '{escape_ts_string(flavor)}'")
    parts.append(f"cardText: '{escape_ts_string(card_text)}'")
    parts.append(f"collectible: true")
    parts.append(f"set: 'CORE'")

    return f"  {{ {', '.join(parts)} }}"


def card_to_art_json(card, faction_name, card_num, card_type):
    """Convert card to art-export JSON format."""
    prefix = FACTION_MAP[faction_name]["prefix"]
    cid = make_id(prefix, card_num, str(card.get("Type", "Minion")))
    kw_str = card.get("Keywords") or ""
    keywords_parsed = parse_keywords(kw_str)
    kw_list = [k[0] for k in keywords_parsed]

    # Detect DEPLOY/LAST_WORDS from card text
    text_upper = (str(card.get("Card Text", "")) or "").upper()
    if ("DEPLOY:" in text_upper or "DEPLOY." in text_upper) and "DEPLOY" not in kw_list:
        kw_list.append("DEPLOY")
    if "LAST WORDS:" in text_upper and "LAST WORDS" not in kw_list:
        kw_list.append("LAST WORDS")

    return {
        "id": cid,
        "name": str(card.get("Card Name", "Unknown")),
        "faction": faction_name,
        "type": str(card.get("Type", "Minion")),
        "tribe": str(card.get("Subtype")) if card.get("Subtype") else None,
        "rarity": str(card.get("Rarity", "Common")),
        "cost": int(card.get("Cost", 0)),
        "attack": int(card["Atk"]) if card.get("Atk") is not None else None,
        "health": int(card["HP"]) if card.get("HP") is not None else None,
        "keywords": kw_list,
        "cardText": str(card.get("Card Text", "") or ""),
        "flavorText": str(card.get("Flavor Text", "") or ""),
        "collectible": True,
        "set": "CORE",
        "artPrompt": "",
    }


def build_starter_deck_ts(faction_name, cards, faction_info):
    """Build a 30-card starter deck from the 'Starter Deck (25)' sheet or auto-pick."""
    # We want 30 cards with good curve
    # Pick: 5 legendaries/epics, 10 rares, 15 commons — with mana curve
    deck_cards = []

    legendaries = [c for c in cards if str(c.get("Rarity","")).upper() == "LEGENDARY"]
    epics = [c for c in cards if str(c.get("Rarity","")).upper() == "EPIC"]
    rares = [c for c in cards if str(c.get("Rarity","")).upper() == "RARE"]
    commons = [c for c in cards if str(c.get("Rarity","")).upper() == "COMMON"]

    # Include up to 2 legendaries, 3 epics
    deck_cards.extend(legendaries[:2])
    deck_cards.extend(epics[:3])

    # Fill with rares (good curve)
    remaining = 30 - len(deck_cards)
    rares_sorted = sorted(rares, key=lambda c: int(c.get("Cost", 0)))
    deck_cards.extend(rares_sorted[:min(10, remaining)])

    remaining = 30 - len(deck_cards)
    commons_sorted = sorted(commons, key=lambda c: int(c.get("Cost", 0)))
    deck_cards.extend(commons_sorted[:remaining])

    deck_cards = deck_cards[:30]
    deck_cards.sort(key=lambda c: (int(c.get("Cost", 0)), str(c.get("Card Name", ""))))

    return deck_cards


def main():
    all_factions = {}
    all_art_cards = []

    # Load all Excel files
    for fname in sorted(os.listdir(CARDS_DIR)):
        if not fname.endswith(".xlsx"):
            continue
        faction_name = fname.replace("STARFORGE_", "").replace("_v2.xlsx", "").replace("_v3.xlsx", "")
        if faction_name not in FACTION_MAP:
            print(f"  WARNING: Unknown faction '{faction_name}', skipping")
            continue

        path = os.path.join(CARDS_DIR, fname)
        sheets = load_excel(path)

        # Get full collection
        full_sheet = None
        for sname, cards in sheets.items():
            if "Full" in sname or "Collection" in sname or len(cards) >= 50:
                full_sheet = cards
                break
        if not full_sheet:
            full_sheet = list(sheets.values())[0]

        # Get starter deck sheet
        starter_sheet = None
        for sname, cards in sheets.items():
            if "Starter" in sname:
                starter_sheet = cards
                break

        all_factions[faction_name] = {
            "cards": full_sheet,
            "starter": starter_sheet,
            "info": FACTION_MAP[faction_name],
        }
        print(f"  Loaded {faction_name}: {len(full_sheet)} cards" +
              (f" + {len(starter_sheet)} starter" if starter_sheet else ""))

    # ================================================================
    # Generate SampleCards.ts
    # ================================================================
    print("\nGenerating SampleCards.ts...")

    ts_lines = []
    ts_lines.append("/**")
    ts_lines.append(" * STARFORGE TCG - Complete Card Collection v2")
    ts_lines.append(" *")
    ts_lines.append(" * Auto-generated from v2 Excel card files.")
    ts_lines.append(f" * {sum(len(f['cards']) for f in all_factions.values())} cards across {len(all_factions)} factions.")
    ts_lines.append(" * Keywords: 8 Combat + 2 Triggers + 11 Original = 21 total")
    ts_lines.append(" *")
    ts_lines.append(" * DO NOT EDIT MANUALLY - regenerate with: python3 scripts/convert-v2-cards.py")
    ts_lines.append(" */")
    ts_lines.append("")
    ts_lines.append("import { CardType, CardRarity, MinionTribe } from '../types/Card';")
    ts_lines.append("import type { CardDefinition } from '../types/Card';")
    ts_lines.append("import { Race } from '../types/Race';")
    ts_lines.append("import { CombatKeyword, TriggerKeyword, OriginalKeyword } from '../types/Keywords';")
    ts_lines.append("import { StarforgeType, StarforgeConditionType } from '../types/Starforge';")
    ts_lines.append("import { EffectType, EffectTrigger, TargetType } from '../types/Effects';")
    ts_lines.append("import type { DamageEffectData, HealEffectData, BuffEffectData, DrawEffectData, SummonEffectData, GrantKeywordData, GainCrystalsData, GenericEffectData, AdaptData, ScryData } from '../types/Effects';")
    ts_lines.append("")

    # Faction card arrays
    for faction_name in ["Cogsmiths", "Luminar", "Pyroclast", "Voidborn", "Biotitans",
                         "Crystalline", "PhantomCorsairs", "Hivemind", "Astromancers", "Chronobound"]:
        if faction_name not in all_factions:
            continue
        fdata = all_factions[faction_name]
        info = fdata["info"]
        cards = fdata["cards"]
        enum_name = info["enum"]

        ts_lines.append(f"// ============================================================================")
        ts_lines.append(f"// {enum_name} ({len(cards)} cards)")
        ts_lines.append(f"// ============================================================================")
        ts_lines.append(f"export const {enum_name}_CARDS: CardDefinition[] = [")

        counters = {"m": 0, "s": 0, "st": 0}
        for i, card in enumerate(cards):
            ct = str(card.get("Type", "Minion"))
            type_code = {"Minion": "m", "Spell": "s", "Structure": "st"}.get(ct, "c")
            counters[type_code] = counters.get(type_code, 0) + 1
            card_num = counters[type_code]
            ts_line = card_to_ts(card, info, card_num, counters)
            ts_lines.append(ts_line + ",")

        ts_lines.append("];")
        ts_lines.append("")

    # Empty NEUTRAL_CARDS (v2 doesn't have neutral)
    ts_lines.append("// ============================================================================")
    ts_lines.append("// NEUTRAL (placeholder - v2 uses faction-only cards)")
    ts_lines.append("// ============================================================================")
    ts_lines.append("export const NEUTRAL_CARDS: CardDefinition[] = [];")
    ts_lines.append("")

    # TOKEN_CARDS
    ts_lines.append("// ============================================================================")
    ts_lines.append("// TOKEN CARDS")
    ts_lines.append("// ============================================================================")
    ts_lines.append("export const TOKEN_CARDS: CardDefinition[] = [];")
    ts_lines.append("")

    # Starter decks
    for faction_name in ["Cogsmiths", "Luminar", "Pyroclast", "Voidborn", "Biotitans",
                         "Crystalline", "PhantomCorsairs", "Hivemind", "Astromancers", "Chronobound"]:
        if faction_name not in all_factions:
            continue
        fdata = all_factions[faction_name]
        info = fdata["info"]
        enum_name = info["enum"]

        # Use starter sheet if available, otherwise auto-build
        if fdata["starter"]:
            starter_cards = fdata["starter"]
        else:
            starter_cards = build_starter_deck_ts(faction_name, fdata["cards"], info)

        display_enum = enum_name.replace("PHANTOM_CORSAIRS", "PHANTOM_CORSAIRS")
        ts_lines.append(f"// Starter Deck: {faction_name} ({len(starter_cards)} cards)")
        ts_lines.append(f"export const STARTER_DECK_{enum_name}: CardDefinition[] = [")

        counters = {"m": 0, "s": 0, "st": 0}
        for i, card in enumerate(starter_cards):
            ct = str(card.get("Type", "Minion"))
            type_code = {"Minion": "m", "Spell": "s", "Structure": "st"}.get(ct, "c")
            counters[type_code] = counters.get(type_code, 0) + 1
            card_num = counters[type_code]
            # Use starter-specific IDs
            orig_prefix = info["prefix"]
            info_copy = dict(info)
            info_copy["prefix"] = f"{orig_prefix}_sd"
            ts_line = card_to_ts(card, info_copy, card_num, counters)
            info["prefix"] = orig_prefix  # restore
            ts_lines.append(ts_line + ",")

        ts_lines.append("];")
        ts_lines.append("")

    # ALL_SAMPLE_CARDS
    ts_lines.append("// ============================================================================")
    ts_lines.append("// COMBINED EXPORTS")
    ts_lines.append("// ============================================================================")
    ts_lines.append("export const ALL_SAMPLE_CARDS: CardDefinition[] = [")
    for faction_name in ["Cogsmiths", "Luminar", "Pyroclast", "Voidborn", "Biotitans",
                         "Crystalline", "PhantomCorsairs", "Hivemind", "Astromancers", "Chronobound"]:
        if faction_name not in all_factions:
            continue
        enum_name = FACTION_MAP[faction_name]["enum"]
        ts_lines.append(f"  ...{enum_name}_CARDS,")
    ts_lines.append("  ...NEUTRAL_CARDS,")
    ts_lines.append("];")
    ts_lines.append("")

    # Helper functions
    ts_lines.append("export function getStarterDeck(race: Race): CardDefinition[] {")
    ts_lines.append("  switch (race) {")
    for faction_name in ["Cogsmiths", "Luminar", "Pyroclast", "Voidborn", "Biotitans",
                         "Crystalline", "PhantomCorsairs", "Hivemind", "Astromancers", "Chronobound"]:
        if faction_name not in all_factions:
            continue
        info = FACTION_MAP[faction_name]
        ts_lines.append(f"    case {info['race']}: return STARTER_DECK_{info['enum']};")
    ts_lines.append("    default: return STARTER_DECK_COGSMITHS;")
    ts_lines.append("  }")
    ts_lines.append("}")
    ts_lines.append("")

    ts_lines.append("export function getSampleCardsByRace(race: Race): CardDefinition[] {")
    ts_lines.append("  switch (race) {")
    for faction_name in ["Cogsmiths", "Luminar", "Pyroclast", "Voidborn", "Biotitans",
                         "Crystalline", "PhantomCorsairs", "Hivemind", "Astromancers", "Chronobound"]:
        if faction_name not in all_factions:
            continue
        info = FACTION_MAP[faction_name]
        ts_lines.append(f"    case {info['race']}: return {info['enum']}_CARDS;")
    ts_lines.append(f"    case Race.NEUTRAL: return NEUTRAL_CARDS;")
    ts_lines.append("    default: return [];")
    ts_lines.append("  }")
    ts_lines.append("}")
    ts_lines.append("")

    ts_lines.append("export function getCollectibleSampleCards(): CardDefinition[] {")
    ts_lines.append("  return ALL_SAMPLE_CARDS.filter(c => c.collectible);")
    ts_lines.append("}")
    ts_lines.append("")

    ts_lines.append("export function getFactionDeck(race: Race): CardDefinition[] {")
    ts_lines.append("  const factionCards = getSampleCardsByRace(race);")
    ts_lines.append("  return factionCards.slice(0, 30);")
    ts_lines.append("}")
    ts_lines.append("")

    # Write SampleCards.ts
    sample_path = os.path.join(SRC_DIR, "SampleCards.ts")
    with open(sample_path, "w") as f:
        f.write("\n".join(ts_lines))
    print(f"  Wrote {sample_path} ({len(ts_lines)} lines)")

    # ================================================================
    # Generate art-export JSON files
    # ================================================================
    print("\nGenerating art-export JSON files...")

    for faction_name, fdata in all_factions.items():
        cards = fdata["cards"]
        info = fdata["info"]
        art_cards = []

        counters = {"m": 0, "s": 0, "st": 0}
        for card in cards:
            ct = str(card.get("Type", "Minion"))
            type_code = {"Minion": "m", "Spell": "s", "Structure": "st"}.get(ct, "c")
            counters[type_code] = counters.get(type_code, 0) + 1
            card_num = counters[type_code]
            art_card = card_to_art_json(card, faction_name, card_num, ct)
            art_cards.append(art_card)
            all_art_cards.append(art_card)

        # Write per-faction JSON
        faction_dir = os.path.join(ART_DIR, faction_name)
        os.makedirs(faction_dir, exist_ok=True)
        art_path = os.path.join(faction_dir, "cards.json")
        with open(art_path, "w") as f:
            json.dump(art_cards, f, indent=2)
        print(f"  Wrote {art_path} ({len(art_cards)} cards)")

    # Write combined JSON
    all_path = os.path.join(ART_DIR, "all-cards.json")
    with open(all_path, "w") as f:
        json.dump(all_art_cards, f, indent=2)
    print(f"  Wrote {all_path} ({len(all_art_cards)} cards total)")

    print(f"\nDone! Updated {len(all_factions)} factions, {len(all_art_cards)} total cards.")


if __name__ == "__main__":
    main()
